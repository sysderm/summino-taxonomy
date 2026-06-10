#!/usr/bin/env python3
"""Harvest the Scopus ASJC (All Science Journal Classification).

Source: Elsevier support article a_id/15181, which links a public XLSX
(ASJC1.xlsx) containing several classifications across sheets. We read only the
"ASJC" sheet: 27 two-digit subject areas + 334 four-digit subject categories.

A four-digit code NNxx is parented to the two-digit subject area NN (its first
two digits). The four super-groups (Life/Physical/Health/Social Sciences) are
not part of this sheet, so subject areas have no parent here.
"""
from __future__ import annotations

import io

import openpyxl

import _common as c

PAGE = "https://service.elsevier.com/app/answers/detail/a_id/15181/supporthub/scopus/"
XLSX = "https://supportcontent.elsevier.com/RightNow%20Next%20Gen/SciVal/ASJC1.xlsx"


def parse(logger) -> None:
    with c.make_client() as client:
        resp = c.fetch(client, XLSX, logger)
    logger.info("fetched %s (%d bytes)", XLSX, len(resp.content))

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True,
                                data_only=True)
    if "ASJC" not in wb.sheetnames:
        raise RuntimeError(f"no ASJC sheet; got {wb.sheetnames}")
    ws = wb["ASJC"]

    nodes: list[dict] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        system, code, desc = row[0], row[1], row[2]
        if code is None or desc is None:
            continue
        code = str(code).strip()
        name = str(desc).strip()
        if not code.isdigit():
            continue
        if code in seen:
            continue
        seen.add(code)
        if len(code) <= 2:  # subject area
            nodes.append(c.normalize_node(
                id=code, name=name, parents=[],
                extras={"level": 1, "kind": "subject_area"}))
        elif len(code) == 4:  # subject category
            area = code[:2]
            nodes.append(c.normalize_node(
                id=code, name=name, parents=[area],
                extras={"level": 2, "kind": "subject_category",
                        "subject_area": area}))
        else:
            logger.warning("unexpected code length: %s", code)

    # Validate every category's parent area exists.
    ids = {n["id"] for n in nodes}
    orphans = [n["id"] for n in nodes
               if n["parents"] and n["parents"][0] not in ids]
    if orphans:
        logger.warning("%d categories with missing parent area: %s",
                       len(orphans), orphans[:10])

    areas = sum(1 for n in nodes if n["extras"]["kind"] == "subject_area")
    cats = sum(1 for n in nodes if n["extras"]["kind"] == "subject_category")
    if cats < 300:
        raise RuntimeError(f"only {cats} categories parsed; expected ~334")
    logger.info("subject_areas=%d subject_categories=%d", areas, cats)
    c.write_raw("scopus-asjc", "ASJC1.xlsx (Elsevier a_id/15181)", nodes, logger)


if __name__ == "__main__":
    raise SystemExit(c.run_parser("scopus-asjc", parse))
